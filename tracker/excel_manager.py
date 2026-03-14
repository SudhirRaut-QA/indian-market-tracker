"""
Excel Manager - Production-Grade Market Data Logger
=====================================================

Reliable, structured Excel logging with:
- Dashboard: Live summary (OVERWRITTEN each run)
- FII_DII: Daily FII/DII with date-based dedup
- Indices: Intraday indices with value-based dedup + Breadth Ratio
- Sectors: Sector summary with 13 columns (incl. 52H/L counts, 30d/1Y)
- Stocks: Per-stock data from all tracked sectors
- Commodities: Gold, Silver, etc.
- Forex: Currency pairs
- Corporate: Full-detail actions with type classification + yield
- Alerts_52W: Near breakout/breakdown alerts
- Insider: Insider trading log with qty + value
- Options: PCR data
- PreOpen: Pre-market data
- BulkBlock: Bulk/Block deals

Deduplication strategy:
- Time-series (Indices, Sectors, Stocks, Commodities, Forex, Options, PreOpen):
  Skip if core values haven't changed since last entry.
- Daily data (FII_DII): Skip if same API date + same net values.
- Event data (Corporate, Insider, BulkBlock): Composite key dedup.
"""

import logging
import os
import re
from datetime import datetime
from typing import Dict, Any, Optional, List, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config

logger = logging.getLogger(__name__)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Styles
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, color="666666", size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Dashboard: which indices to show in summary
DASHBOARD_INDICES = [
    "NIFTY 50", "NIFTY BANK", "INDIA VIX", "NIFTY IT",
    "NIFTY PHARMA", "NIFTY METAL", "NIFTY MIDCAP 50",
]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Helpers
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _get_wb(path: str) -> Workbook:
    """Load existing workbook or create new one."""
    if os.path.exists(path):
        return load_workbook(path)
    return Workbook()


def _header_row(ws, cols: List[str], row: int = 1):
    """Write styled header row."""
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)


def _write_row(ws, row: int, vals: list, color_cols: dict = None):
    """Write a row of values with borders and optional conditional coloring.

    color_cols: dict mapping 1-based column index to 'green_red' rule.
    """
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = THIN_BORDER
        if color_cols and i in color_cols:
            if isinstance(v, (int, float)):
                if v > 0:
                    cell.fill = GREEN_FILL
                elif v < 0:
                    cell.fill = RED_FILL


def _classify_action(subject: str) -> str:
    """Classify corporate action type from subject text."""
    s = subject.lower()
    if any(k in s for k in ("dividend", "interim div", "final div")):
        return "Dividend"
    if "bonus" in s:
        return "Bonus"
    if any(k in s for k in ("split", "sub-division", "subdivision")):
        return "Split"
    if "rights" in s:
        return "Rights"
    if "interest" in s:
        return "Interest"
    if any(k in s for k in ("buyback", "buy back", "buy-back")):
        return "Buyback"
    return "Other"


def _extract_dividend_amount(subject: str) -> float:
    """Extract dividend amount from subject like 'Rs 10 Per Share'."""
    patterns = [
        r'Rs\.?\s*([\d.]+)\s*(?:/|-|per)\s*(?:share|shr)',
        r'Re\.?\s*([\d.]+)\s*(?:/|-|per)\s*(?:share|shr)',
        r'Rs\.?\s*([\d.]+)',
        r'Re\.?\s*([\d.]+)',
    ]
    for pat in patterns:
        m = re.search(pat, subject, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                pass
    return 0.0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Excel Manager
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class ExcelManager:
    """Production-grade Excel manager with deduplication and structured sheets."""

    SHEET_ORDER = [
        "Dashboard", "Trading", "FII_DII", "Indices", "Sectors", "Stocks",
        "Commodities", "Forex", "Corporate", "Alerts_52W",
        "Insider", "Options", "PreOpen", "BulkBlock",
    ]

    def __init__(self, path: str = None):
        self.path = path or str(config.EXCEL_FILE)
        os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
        os.makedirs(config.EXCEL_DIR, exist_ok=True)

    # ── Main Entry Point ─────────────────────────────────────────────────────

    def log_snapshot(self, snapshot: Dict, delta: Optional[Dict] = None,
                     trading_setups: Optional[Dict] = None):
        """Log all data from snapshot to Excel with deduplication."""
        try:
            wb = _get_wb(self.path)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            date = datetime.now().strftime("%Y-%m-%d")

            # Dashboard: OVERWRITE with latest summary
            self._update_dashboard(wb, snapshot, ts, trading_setups)

            # Historical sheets: APPEND with dedup
            self._log_fii_dii(wb, snapshot, ts)
            self._log_indices(wb, snapshot, ts)
            self._log_sectors(wb, snapshot, ts)
            self._log_stocks(wb, snapshot, ts)
            self._log_commodities(wb, snapshot, ts)
            self._log_forex(wb, snapshot, ts)
            self._log_options(wb, snapshot, ts)
            self._log_preopen(wb, snapshot, ts)

            if snapshot.get("corporate_actions"):
                self._log_corporate(wb, snapshot, date)
            if snapshot.get("insider_trading"):
                self._log_insider(wb, snapshot, date)
            if snapshot.get("bulk_deals") or snapshot.get("block_deals"):
                self._log_bulk_deals(wb, snapshot, date)
            if snapshot.get("sectors"):
                self._log_alerts(wb, snapshot, ts)

            # Trading setups sheet
            if trading_setups:
                self._log_trading(wb, trading_setups, ts)

            # Remove default "Sheet" if empty
            if "Sheet" in wb.sheetnames:
                ws_default = wb["Sheet"]
                if ws_default.max_row == 1 and ws_default.max_column == 1 and not ws_default.cell(1, 1).value:
                    wb.remove(ws_default)

            # Reorder tabs
            self._reorder_sheets(wb)

            wb.save(self.path)
            logger.info(f"Excel saved: {self.path}")
        except PermissionError:
            logger.error(f"Excel file is open in another app — close it and retry: {self.path}")
        except Exception as e:
            logger.error(f"Excel save error: {e}")

    # ── Sheet Helpers ────────────────────────────────────────────────────────

    def _get_or_create_sheet(self, wb: Workbook, name: str, headers: List[str]):
        """Get existing sheet or create fresh one with headers.

        If headers mismatch (column count or names), backs up old sheet
        as '<name>_OLD' and creates a new one for data integrity.
        """
        if name in wb.sheetnames:
            ws = wb[name]
            # Validate headers match
            existing_headers = []
            for i in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=i).value
                if val is not None:
                    existing_headers.append(val)
            if existing_headers == headers:
                return ws
            # Headers mismatch — back up and recreate
            old_name = f"{name}_OLD"
            if old_name in wb.sheetnames:
                del wb[old_name]
            ws.title = old_name
            logger.warning(
                f"Sheet '{name}': headers changed "
                f"({len(existing_headers)} cols -> {len(headers)} cols). "
                f"Old data backed up to '{old_name}'."
            )

        ws = wb.create_sheet(name)
        _header_row(ws, headers)
        return ws

    def _reorder_sheets(self, wb: Workbook):
        """Reorder sheet tabs to match SHEET_ORDER."""
        desired = [n for n in self.SHEET_ORDER if n in wb.sheetnames]
        extra = [n for n in wb.sheetnames if n not in desired]
        desired.extend(extra)
        wb._sheets = [wb[n] for n in desired]

    # ── Dashboard (OVERWRITE each run) ───────────────────────────────────────

    def _update_dashboard(self, wb: Workbook, snapshot: Dict, ts: str,
                          trading_setups: Optional[Dict] = None):
        """Overwrite Dashboard with latest market summary."""
        name = "Dashboard"
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name, 0)

        row = 1
        ws.cell(row=row, column=1, value="INDIAN MARKET TRACKER").font = TITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=f"Last Updated: {ts} IST").font = SUBTITLE_FONT
        row += 2

        # ── Key Indices ──
        ws.cell(row=row, column=1, value="Key Indices").font = SECTION_FONT
        row += 1
        _header_row(ws, ["Index", "Last", "Change", "% Change", "Adv", "Dec"], row)
        row += 1
        indices = snapshot.get("indices") or {}
        for idx_name in DASHBOARD_INDICES:
            data = indices.get(idx_name)
            if not data:
                continue
            display = idx_name.replace("NIFTY ", "").replace("INDIA ", "")
            vals = [
                display, data.get("last", 0), data.get("change", 0),
                data.get("pct", 0), data.get("advances", 0), data.get("declines", 0),
            ]
            _write_row(ws, row, vals, {3: "green_red", 4: "green_red"})
            row += 1
        row += 1

        # ── Commodities & Forex ──
        ws.cell(row=row, column=1, value="Commodities & Forex").font = SECTION_FONT
        row += 1
        comms = snapshot.get("commodities") or {}
        for sym, data in comms.items():
            pct = data.get("pct", 0)
            pct_str = f"{pct:.2f}%" if isinstance(pct, (int, float)) else str(pct)
            vals = [
                sym, data.get("last", 0), pct_str,
                f"52H: {data.get('week52_high', 0)}",
                f"52L: {data.get('week52_low', 0)}",
            ]
            _write_row(ws, row, vals)
            row += 1
        forex = snapshot.get("forex") or {}
        if forex:
            _write_row(ws, row, [
                "USD/INR", forex.get("usdinr", 0), forex.get("date", ""),
            ])
            row += 1
        row += 1

        # ── FII/DII Summary ──
        fd = snapshot.get("fii_dii")
        if fd:
            ws.cell(row=row, column=1, value="FII/DII Activity").font = SECTION_FONT
            ws.cell(row=row, column=4, value=f"Date: {fd.get('date', '')}").font = SUBTITLE_FONT
            row += 1
            _header_row(ws, ["Category", "Buy (Cr)", "Sell (Cr)", "Net (Cr)"], row)
            row += 1
            _write_row(ws, row, [
                "FII/FPI",
                fd.get("fii", {}).get("buy", 0),
                fd.get("fii", {}).get("sell", 0),
                fd.get("fii", {}).get("net", 0),
            ], {4: "green_red"})
            row += 1
            _write_row(ws, row, [
                "DII",
                fd.get("dii", {}).get("buy", 0),
                fd.get("dii", {}).get("sell", 0),
                fd.get("dii", {}).get("net", 0),
            ], {4: "green_red"})
            row += 1
            _write_row(ws, row, [
                f"Signal: {fd.get('signal', '')}",
                "", "", fd.get("interpretation", ""),
            ])
            row += 2

        # ── 52-Week Alerts ──
        alerts = self._collect_52w_alerts(snapshot)
        if alerts:
            ws.cell(row=row, column=1, value="52-Week Alerts").font = SECTION_FONT
            row += 1
            _header_row(ws, ["Alert", "Symbol", "LTP", "52W Level", "Distance"], row)
            row += 1
            for a in alerts[:15]:
                _write_row(ws, row, [
                    a["type"], a["symbol"], a["ltp"],
                    a["level"], f"{a['distance']:.2f}% away",
                ])
                cell = ws.cell(row=row, column=1)
                cell.fill = GREEN_FILL if "HIGH" in a["type"] else RED_FILL
                row += 1

        # ── Intraday Trading Hints ──
        if trading_setups:
            row += 1
            bias = trading_setups.get("bias", {})
            ws.cell(row=row, column=1, value="Intraday Trading Analysis").font = SECTION_FONT
            row += 1
            bias_dir = bias.get("direction", "NEUTRAL")
            bias_score = bias.get("score", 0)
            ws.cell(row=row, column=1, value=f"Market Bias: {bias_dir} (score {bias_score:+d}/100)")
            fill = GREEN_FILL if bias_dir == "BULLISH" else (RED_FILL if bias_dir == "BEARISH" else YELLOW_FILL)
            ws.cell(row=row, column=1).fill = fill
            row += 1
            for r_text in bias.get("reasons", [])[:4]:
                ws.cell(row=row, column=1, value=f"  • {r_text}")
                row += 1
            row += 1

            # Index levels summary
            idx_setups = trading_setups.get("index_setups", [])
            if idx_setups:
                _header_row(ws, ["Index", "LTP", "Pivot", "S1", "S2", "R1", "R2", "Direction", "Entry", "Target", "SL", "R:R"], row)
                row += 1
                for s in idx_setups:
                    vals = [
                        s["symbol"], s["ltp"], s["classic_pivot"],
                        s["classic_s1"], s["classic_s2"],
                        s["classic_r1"], s["classic_r2"],
                        s["direction"], s["entry"], s["target"],
                        s["stop_loss"], s["risk_reward"],
                    ]
                    _write_row(ws, row, vals)
                    row += 1
                row += 1

            # Top stock setups
            stk_setups = trading_setups.get("stock_setups", [])
            top_stk = [s for s in stk_setups if s["direction"] != "NEUTRAL"][:10]
            if top_stk:
                ws.cell(row=row, column=1, value="Top Stock Setups").font = SECTION_FONT
                row += 1
                _header_row(ws, ["Symbol", "Sector", "LTP", "Direction", "Entry", "Target", "SL", "R:R", "Key Factor"], row)
                row += 1
                for s in top_stk:
                    factor = s["factors"][0] if s.get("factors") else ""
                    vals = [
                        s["symbol"], s["sector"], s["ltp"],
                        s["direction"], s["entry"], s["target"],
                        s["stop_loss"], s["risk_reward"], factor,
                    ]
                    _write_row(ws, row, vals)
                    cell = ws.cell(row=row, column=4)
                    cell.fill = GREEN_FILL if s["direction"] == "LONG" else RED_FILL
                    row += 1

        _auto_width(ws)

    # ── FII/DII (Append, dedup by API date + net) ───────────────────────────

    def _log_fii_dii(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "FII_DII"
        fd = snapshot.get("fii_dii")
        if not fd:
            return

        headers = [
            "Timestamp", "Date",
            "FII Buy (Cr)", "FII Sell (Cr)", "FII Net (Cr)",
            "DII Buy (Cr)", "DII Sell (Cr)", "DII Net (Cr)",
            "Total Net (Cr)", "Signal", "Interpretation",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: skip if last row has same FII date AND same FII net value
        fii_date = fd.get("date", "")
        fii_net = fd["fii"]["net"]
        if ws.max_row >= 2:
            last_date = ws.cell(row=ws.max_row, column=2).value
            last_fii_net = ws.cell(row=ws.max_row, column=5).value
            if last_date == fii_date and last_fii_net == fii_net:
                logger.debug(f"FII_DII: skipped duplicate ({fii_date})")
                return

        row = ws.max_row + 1
        vals = [
            ts, fii_date,
            fd["fii"]["buy"], fd["fii"]["sell"], fd["fii"]["net"],
            fd["dii"]["buy"], fd["dii"]["sell"], fd["dii"]["net"],
            fd.get("total_net", 0), fd.get("signal", ""), fd.get("interpretation", ""),
        ]
        _write_row(ws, row, vals, {5: "green_red", 8: "green_red", 9: "green_red"})
        _auto_width(ws)

    # ── Indices (Append, value-based dedup) ──────────────────────────────────

    def _log_indices(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Indices"
        indices = snapshot.get("indices")
        if not indices:
            return

        headers = [
            "Timestamp", "Index", "Last", "Change", "% Change",
            "Open", "High", "Low", "Prev Close",
            "Advances", "Declines", "Breadth Ratio",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: skip if NIFTY 50 last price is unchanged
        nifty = indices.get("NIFTY 50", {})
        if nifty and ws.max_row >= 2:
            for r in range(ws.max_row, max(1, ws.max_row - 100), -1):
                if ws.cell(row=r, column=2).value == "NIFTY 50":
                    if ws.cell(row=r, column=3).value == nifty.get("last"):
                        logger.debug("Indices: skipped (values unchanged)")
                        return
                    break

        for idx_name, data in indices.items():
            row = ws.max_row + 1
            try:
                adv = int(data.get("advances", 0) or 0)
                dec = int(data.get("declines", 0) or 0)
            except (ValueError, TypeError):
                adv, dec = 0, 0
            breadth = round(adv / dec, 2) if dec > 0 else 0
            vals = [
                ts, idx_name, data["last"], data["change"], data["pct"],
                data["open"], data["high"], data["low"], data["prev_close"],
                adv, dec, breadth,
            ]
            _write_row(ws, row, vals, {5: "green_red"})

    # ── Sectors (Append with full metrics) ───────────────────────────────────

    def _log_sectors(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Sectors"
        sectors = snapshot.get("sectors")
        if not sectors:
            return

        headers = [
            "Timestamp", "Sector", "Index Last", "Index %",
            "Stocks", "Top Gainer", "G %", "Top Loser", "L %",
            "Near 52H Count", "Near 52L Count", "30d Chg (Avg)", "1Y Chg (Avg)",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check first sector's index value
        first_sector = next(iter(sectors.values()), {})
        if first_sector and ws.max_row >= 2:
            sect_name_check = first_sector.get("sector", "")
            for r in range(ws.max_row, max(1, ws.max_row - 100), -1):
                if ws.cell(row=r, column=2).value == sect_name_check:
                    if ws.cell(row=r, column=3).value == first_sector.get("index_last"):
                        logger.debug("Sectors: skipped (values unchanged)")
                        return
                    break

        for sect_name, data in sectors.items():
            row = ws.max_row + 1
            stocks = data.get("stocks", [])
            g = (data.get("gainers") or [{}])[0] if data.get("gainers") else {}
            lo = (data.get("losers") or [{}])[0] if data.get("losers") else {}

            near_52h = sum(
                1 for s in stocks
                if isinstance(s.get("near_52h"), (int, float)) and 0 < s["near_52h"] <= config.NEAR_52W_HIGH_PCT
            )
            near_52l = sum(
                1 for s in stocks
                if isinstance(s.get("near_52l"), (int, float)) and 0 < s["near_52l"] <= config.NEAR_52W_LOW_PCT
            )
            chg30 = [s.get("chg_30d", 0) for s in stocks if isinstance(s.get("chg_30d"), (int, float))]
            chg1y = [s.get("chg_365d", 0) for s in stocks if isinstance(s.get("chg_365d"), (int, float))]
            avg_30d = round(sum(chg30) / len(chg30), 2) if chg30 else 0
            avg_1y = round(sum(chg1y) / len(chg1y), 2) if chg1y else 0

            vals = [
                ts, sect_name, data.get("index_last", 0), data.get("index_pct", 0),
                data.get("count", 0), g.get("symbol", ""), g.get("pct", 0),
                lo.get("symbol", ""), lo.get("pct", 0),
                near_52h, near_52l, avg_30d, avg_1y,
            ]
            _write_row(ws, row, vals)

    # ── Stocks (Per-stock data from sectors) ─────────────────────────────────

    def _log_stocks(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Stocks"
        sectors = snapshot.get("sectors")
        if not sectors:
            return

        headers = [
            "Timestamp", "Sector", "Symbol", "Last", "Change", "% Change",
            "Volume", "Value (Cr)", "52W High", "52W Low",
            "Near 52H %", "Near 52L %", "30d %", "1Y %",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check lead stock of first sector
        first_sect = next(iter(sectors.values()), {})
        first_stocks = first_sect.get("stocks", [])
        if first_stocks and ws.max_row >= 2:
            check_sym = first_stocks[0].get("symbol", "")
            check_ltp = first_stocks[0].get("last")
            for r in range(ws.max_row, max(1, ws.max_row - 600), -1):
                if ws.cell(row=r, column=3).value == check_sym:
                    if ws.cell(row=r, column=4).value == check_ltp:
                        logger.debug("Stocks: skipped (values unchanged)")
                        return
                    break

        for sect_name, data in sectors.items():
            for s in data.get("stocks", []):
                row = ws.max_row + 1
                vals = [
                    ts, sect_name, s.get("symbol", ""),
                    s.get("last", 0), s.get("change", 0), s.get("pct", 0),
                    s.get("volume", 0), s.get("value_cr", 0),
                    s.get("year_high", 0), s.get("year_low", 0),
                    s.get("near_52h", 0), s.get("near_52l", 0),
                    s.get("chg_30d", 0), s.get("chg_365d", 0),
                ]
                _write_row(ws, row, vals, {6: "green_red"})

    # ── Commodities ──────────────────────────────────────────────────────────

    def _log_commodities(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Commodities"
        comms = snapshot.get("commodities")
        if not comms:
            return

        headers = [
            "Timestamp", "Symbol", "Last", "Change", "% Change",
            "52W High", "52W Low",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check first commodity LTP
        first_sym = next(iter(comms.keys()), "")
        first_data = comms.get(first_sym, {})
        if first_data and ws.max_row >= 2:
            for r in range(ws.max_row, 1, -1):
                if ws.cell(row=r, column=2).value == first_sym:
                    if ws.cell(row=r, column=3).value == first_data.get("last"):
                        logger.debug("Commodities: skipped (values unchanged)")
                        return
                    break

        for sym, data in comms.items():
            row = ws.max_row + 1
            vals = [
                ts, sym, data["last"], data["change"], data["pct"],
                data.get("week52_high", 0), data.get("week52_low", 0),
            ]
            _write_row(ws, row, vals)

    # ── Forex ────────────────────────────────────────────────────────────────

    def _log_forex(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Forex"
        forex = snapshot.get("forex")
        if not forex:
            return

        headers = ["Timestamp", "USD/INR", "USD/EUR", "USD/GBP", "USD/JPY", "API Date"]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check USD/INR value
        if ws.max_row >= 2:
            if ws.cell(row=ws.max_row, column=2).value == forex.get("usdinr"):
                logger.debug("Forex: skipped (values unchanged)")
                return

        row = ws.max_row + 1
        vals = [
            ts, forex.get("usdinr", 0), forex.get("usdeur", 0),
            forex.get("usdgbp", 0), forex.get("usdjpy", 0), forex.get("date", ""),
        ]
        _write_row(ws, row, vals)

    # ── Corporate Actions (full columns + composite key dedup) ───────────────

    def _log_corporate(self, wb: Workbook, snapshot: Dict, date: str):
        name = "Corporate"
        actions = snapshot.get("corporate_actions", [])
        if not actions:
            return

        headers = [
            "Log Date", "Type", "Symbol", "Company", "Subject",
            "Ex-Date", "Record Date", "BC Start", "BC End",
            "LTP", "% Change", "PE", "Div Amount", "Div Yield %",
            "52W High", "52W Low",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Build set of existing (symbol, subject) for dedup
        existing: Set[Tuple] = set()
        for r in range(2, ws.max_row + 1):
            key = (
                ws.cell(row=r, column=3).value,   # Symbol
                ws.cell(row=r, column=5).value,    # Subject
            )
            existing.add(key)

        for a in actions:
            key = (a["symbol"], a["subject"])
            if key in existing:
                continue
            existing.add(key)

            action_type = _classify_action(a.get("subject", ""))
            ltp = a.get("ltp", 0) or 0
            try:
                ltp = float(ltp)
            except (ValueError, TypeError):
                ltp = 0.0
            pe = a.get("pe", 0) or 0
            try:
                pe = float(pe)
            except (ValueError, TypeError):
                pe = 0.0
            pct_change = a.get("pct", 0) or 0

            div_amount = 0.0
            div_yield = 0.0
            if action_type == "Dividend":
                div_amount = _extract_dividend_amount(a.get("subject", ""))
                if div_amount > 0 and ltp > 0:
                    div_yield = round(div_amount / ltp * 100, 2)

            row = ws.max_row + 1
            vals = [
                date, action_type, a["symbol"], a["company"], a["subject"],
                a["ex_date"], a.get("record_date", ""),
                a.get("bc_start", ""), a.get("bc_end", ""),
                ltp, pct_change, pe, div_amount, div_yield,
                a.get("week52_high", 0), a.get("week52_low", 0),
            ]
            _write_row(ws, row, vals)

    # ── 52-Week Alerts ───────────────────────────────────────────────────────

    def _log_alerts(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Alerts_52W"
        alerts = self._collect_52w_alerts(snapshot)
        if not alerts:
            return

        headers = [
            "Timestamp", "Alert Type", "Symbol", "Sector",
            "LTP", "52W Level", "Distance %", "Today %",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check first alert's symbol + LTP
        if alerts and ws.max_row >= 2:
            first = alerts[0]
            for r in range(ws.max_row, max(1, ws.max_row - 50), -1):
                if ws.cell(row=r, column=3).value == first["symbol"]:
                    if ws.cell(row=r, column=5).value == first["ltp"]:
                        logger.debug("Alerts_52W: skipped (values unchanged)")
                        return
                    break

        for a in alerts:
            row = ws.max_row + 1
            vals = [
                ts, a["type"], a["symbol"], a["sector"],
                a["ltp"], a["level"],
                round(a["distance"], 2), a.get("today_pct", 0),
            ]
            _write_row(ws, row, vals)
            cell = ws.cell(row=row, column=2)
            cell.fill = GREEN_FILL if "HIGH" in a["type"] else RED_FILL

    # ── Insider Trading ──────────────────────────────────────────────────────

    def _log_insider(self, wb: Workbook, snapshot: Dict, date: str):
        name = "Insider"
        trades = snapshot.get("insider_trading", [])
        if not trades:
            return

        headers = [
            "Log Date", "Symbol", "Company", "Acquirer", "Relation",
            "Buy Qty", "Sell Qty", "Buy Value (Cr)", "Sell Value (Cr)",
            "Trade Date",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Build existing keys: (symbol, acquirer, trade_date)
        existing: Set[Tuple] = set()
        for r in range(2, ws.max_row + 1):
            key = (
                ws.cell(row=r, column=2).value,    # Symbol
                ws.cell(row=r, column=4).value,    # Acquirer
                ws.cell(row=r, column=10).value,   # Trade Date
            )
            existing.add(key)

        for t in trades[:30]:
            key = (t["symbol"], t["acquirer"], t["date"])
            if key in existing:
                continue
            existing.add(key)

            row = ws.max_row + 1
            vals = [
                date, t["symbol"], t["company"], t["acquirer"],
                t.get("relation", ""),
                t.get("buy_qty", 0), t.get("sell_qty", 0),
                t["buy_value"], t["sell_value"],
                t["date"],
            ]
            _write_row(ws, row, vals)

    # ── Options ──────────────────────────────────────────────────────────────

    def _log_options(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "Options"
        oc = snapshot.get("option_chain")
        if not oc:
            return

        headers = [
            "Timestamp", "Symbol", "PCR (OI)", "PCR (Vol)", "Signal",
            "Max Pain", "CE OI Total", "PE OI Total",
            "Top CE Strike 1", "Top CE OI 1",
            "Top PE Strike 1", "Top PE OI 1",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check first symbol's PCR
        first_sym = next(iter(oc.keys()), "")
        first_data = oc.get(first_sym, {})
        if first_data and ws.max_row >= 2:
            for r in range(ws.max_row, 1, -1):
                if ws.cell(row=r, column=2).value == first_sym:
                    if ws.cell(row=r, column=3).value == first_data.get("pcr_oi"):
                        logger.debug("Options: skipped (values unchanged)")
                        return
                    break

        for sym, data in oc.items():
            row = ws.max_row + 1
            top_ce = data.get("top_ce", [{}])
            top_pe = data.get("top_pe", [{}])
            ce1 = top_ce[0] if top_ce else {}
            pe1 = top_pe[0] if top_pe else {}
            vals = [
                ts, sym,
                data.get("pcr_oi", 0), data.get("pcr_vol", 0),
                data.get("signal", ""),
                data.get("max_pain", 0),
                data.get("ce_oi_total", 0), data.get("pe_oi_total", 0),
                ce1.get("strike", 0), ce1.get("oi", 0),
                pe1.get("strike", 0), pe1.get("oi", 0),
            ]
            _write_row(ws, row, vals)

    # ── Pre-Open ─────────────────────────────────────────────────────────────

    def _log_preopen(self, wb: Workbook, snapshot: Dict, ts: str):
        name = "PreOpen"
        po = snapshot.get("preopen")
        if not po:
            return

        headers = [
            "Timestamp", "Index", "Advances", "Declines",
            "Top Gainer", "G IEP", "G %",
            "Top Loser", "L IEP", "L %",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup
        if ws.max_row >= 2:
            g = (po.get("gainers") or [{}])[0] if po.get("gainers") else {}
            last_adv = ws.cell(row=ws.max_row, column=3).value
            last_gainer = ws.cell(row=ws.max_row, column=5).value
            if last_adv == po.get("advances") and last_gainer == g.get("symbol", ""):
                logger.debug("PreOpen: skipped (values unchanged)")
                return

        row = ws.max_row + 1
        g = (po.get("gainers") or [{}])[0] if po.get("gainers") else {}
        lo = (po.get("losers") or [{}])[0] if po.get("losers") else {}
        vals = [
            ts, po.get("key", "NIFTY"),
            po.get("advances", 0), po.get("declines", 0),
            g.get("symbol", ""), g.get("iep", 0), g.get("pct", 0),
            lo.get("symbol", ""), lo.get("iep", 0), lo.get("pct", 0),
        ]
        _write_row(ws, row, vals)

    # ── Bulk/Block Deals ─────────────────────────────────────────────────────

    def _log_bulk_deals(self, wb: Workbook, snapshot: Dict, date: str):
        name = "BulkBlock"
        bulk = snapshot.get("bulk_deals") or []
        block = snapshot.get("block_deals") or []

        if not bulk and not block:
            return

        headers = [
            "Log Date", "Type", "Symbol", "Client",
            "Trade Type", "Quantity", "Price", "Value (Cr)",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Build existing keys: (type, symbol, client)
        existing: Set[Tuple] = set()
        for r in range(2, ws.max_row + 1):
            key = (
                ws.cell(row=r, column=2).value,
                ws.cell(row=r, column=3).value,
                ws.cell(row=r, column=4).value,
            )
            existing.add(key)

        for d in bulk[:30]:
            key = ("BULK", d["symbol"], d["client"])
            if key in existing:
                continue
            existing.add(key)
            row = ws.max_row + 1
            vals = [
                date, "BULK", d["symbol"], d["client"],
                d["trade_type"], d["qty"], d["price"], d["value_cr"],
            ]
            _write_row(ws, row, vals)
            ws.cell(row=row, column=5).fill = GREEN_FILL if d["trade_type"] == "BUY" else RED_FILL

        for d in block[:30]:
            key = ("BLOCK", d["symbol"], d["client"])
            if key in existing:
                continue
            existing.add(key)
            row = ws.max_row + 1
            vals = [
                date, "BLOCK", d["symbol"], d["client"],
                d["trade_type"], d["qty"], d["price"], d["value_cr"],
            ]
            _write_row(ws, row, vals)
            ws.cell(row=row, column=5).fill = GREEN_FILL if d["trade_type"] == "BUY" else RED_FILL

    # ── Utility: Collect 52W alerts from sector data ─────────────────────────

    def _collect_52w_alerts(self, snapshot: Dict) -> List[Dict]:
        """Scan all sector stocks for near-52W high/low alerts."""
        alerts = []
        sectors = snapshot.get("sectors") or {}
        seen: Set[str] = set()

        for sect_name, data in sectors.items():
            for s in data.get("stocks", []):
                sym = s.get("symbol", "")
                if sym in seen:
                    continue

                near_h = s.get("near_52h")
                near_l = s.get("near_52l")
                ltp = s.get("last", 0)

                if isinstance(near_h, (int, float)) and 0 < near_h <= config.NEAR_52W_HIGH_PCT:
                    seen.add(sym)
                    alerts.append({
                        "type": "NEAR 52W HIGH",
                        "symbol": sym,
                        "sector": sect_name,
                        "ltp": ltp,
                        "level": s.get("year_high", 0),
                        "distance": near_h,
                        "today_pct": s.get("pct", 0),
                    })
                elif isinstance(near_l, (int, float)) and 0 < near_l <= config.NEAR_52W_LOW_PCT:
                    seen.add(sym)
                    alerts.append({
                        "type": "NEAR 52W LOW",
                        "symbol": sym,
                        "sector": sect_name,
                        "ltp": ltp,
                        "level": s.get("year_low", 0),
                        "distance": near_l,
                        "today_pct": s.get("pct", 0),
                    })

        alerts.sort(key=lambda a: a["distance"])
        return alerts

    # ── Trading Setups (Append with dedup) ───────────────────────────────────

    def _log_trading(self, wb: Workbook, setups: Dict, ts: str):
        """Log intraday trading setups to the Trading sheet."""
        name = "Trading"
        headers = [
            "Timestamp", "Category", "Symbol", "Sector", "LTP", "% Chg",
            "Pivot", "S1", "S2", "R1", "R2",
            "CPR Low", "CPR High", "CPR Width%",
            "VWAP", "Direction", "Score",
            "Entry", "Target", "Stop Loss", "R:R",
            "Key Factors",
        ]
        ws = self._get_or_create_sheet(wb, name, headers)

        # Dedup: check timestamp of last row
        if ws.max_row >= 2:
            last_ts = ws.cell(row=ws.max_row, column=1).value
            if last_ts == ts:
                logger.debug("Trading: skipped (same timestamp)")
                return

        all_setups = (
            setups.get("index_setups", [])
            + setups.get("stock_setups", [])
            + setups.get("etf_setups", [])
        )

        for s in all_setups:
            row = ws.max_row + 1
            factors_str = " | ".join(s.get("factors", [])[:3])
            vals = [
                ts, s["category"], s["symbol"], s.get("sector", ""),
                s["ltp"], s.get("pct", 0),
                s["classic_pivot"], s["classic_s1"], s["classic_s2"],
                s["classic_r1"], s["classic_r2"],
                s["cpr_bc"], s["cpr_tc"], s["cpr_width_pct"],
                s["vwap"], s["direction"], s["direction_score"],
                s["entry"], s["target"], s["stop_loss"], s["risk_reward"],
                factors_str,
            ]
            _write_row(ws, row, vals, {6: "green_red", 16: "green_red"})
            # Color the direction column
            cell = ws.cell(row=row, column=16)
            if s["direction"] == "LONG":
                cell.fill = GREEN_FILL
            elif s["direction"] == "SHORT":
                cell.fill = RED_FILL
            elif s["direction"] == "NEUTRAL":
                cell.fill = YELLOW_FILL

        _auto_width(ws)
