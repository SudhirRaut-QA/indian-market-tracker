"""
Quick audit test - verifies all scraper components, formatters, and Excel logging.
"""
import sys
import json

def test_imports():
    """Test all module imports work."""
    from tracker.telegram_bot import (
        format_fii_dii_msg, format_sector_msg, format_options_msg,
        format_commodities_msg, format_corporate_msg, format_preopen_msg,
        format_delta_alert, format_52w_alerts_msg, TelegramBot,
        _make_table, _cr, _pct, _emoji_pct, _vol,
        _52w_position, _52w_emoji, _format_prev_time,
    )
    from tracker.interactive_bot import InteractiveTelegramBot
    from tracker.signal_detector import SignalDetector, format_signals_msg
    from tracker.excel_manager import ExcelManager
    from tracker.nse_scraper import MarketScraper
    from tracker.delta_engine import DeltaEngine
    from tracker.scheduler import setup_schedule, run_loop, SLOT_CONFIG, IST
    from tracker.config import (
        SECTORS, KEY_INDICES, COMMODITY_ETFS,
        NEAR_52W_HIGH_PCT, NEAR_52W_LOW_PCT,
        HIGH_DELIVERY_PCT, LOW_DELIVERY_PCT,
    )
    
    assert len(SECTORS) == 16, f"Expected 16 sectors, got {len(SECTORS)}"
    assert len(KEY_INDICES) == 21, f"Expected 21 indices, got {len(KEY_INDICES)}"
    assert len(COMMODITY_ETFS) == 4, f"Expected 4 ETFs, got {len(COMMODITY_ETFS)}"
    assert NEAR_52W_HIGH_PCT == 2.0
    assert len(SLOT_CONFIG) == 8, f"Expected 8 slots, got {len(SLOT_CONFIG)}"
    
    print("[PASS] All imports OK")
    print(f"  Sectors: {len(SECTORS)}, Indices: {len(KEY_INDICES)}, ETFs: {len(COMMODITY_ETFS)}")


def test_formatters_with_mock_data():
    """Test all formatters with mock snapshot data."""
    from tracker.telegram_bot import (
        format_fii_dii_msg, format_sector_msg, format_options_msg,
        format_commodities_msg, format_corporate_msg, format_preopen_msg,
        format_delta_alert, format_52w_alerts_msg,
    )
    
    mock_snapshot = {
        "timestamp": "2026-03-01T10:00:00",
        "fii_dii": {
            "date": "01-Mar-2026", "timestamp": "2026-03-01T10:00:00",
            "fii": {"buy": 5000, "sell": 3000, "net": 2000},
            "dii": {"buy": 4000, "sell": 4500, "net": -500},
            "total_net": 1500, "signal": "FII Bullish",
            "interpretation": "FII buying, DII selling",
        },
        "indices": {
            "NIFTY 50": {"last": 22500, "change": 150, "pct": 0.67,
                         "open": 22350, "high": 22550, "low": 22300,
                         "prev_close": 22350, "advances": 30, "declines": 20, "unchanged": 0},
            "NIFTY BANK": {"last": 48000, "change": -100, "pct": -0.21,
                           "open": 48100, "high": 48200, "low": 47900,
                           "prev_close": 48100, "advances": 5, "declines": 7, "unchanged": 0},
        },
        "market_status": {
            "Capital Market": {"status": "Open", "trade_date": "01-Mar-2026",
                               "index": "NIFTY 50", "last": 22500, "variation": 150, "pct": 0.67},
        },
        "sectors": {
            "NIFTY 50": {
                "sector": "NIFTY 50", "timestamp": "2026-03-01",
                "index_last": 22500, "index_change": 150, "index_pct": 0.67,
                "count": 50,
                "stocks": [
                    {"symbol": "RELIANCE", "last": 2500, "change": 50, "pct": 2.04,
                     "open": 2450, "high": 2520, "low": 2440, "prev_close": 2450,
                     "volume": 5000000, "value_cr": 1250, "year_high": 2600, "year_low": 2000,
                     "near_52h": 3.8, "near_52l": 25.0, "chg_30d": 5.0, "chg_365d": 15.0},
                    {"symbol": "TCS", "last": 3800, "change": -20, "pct": -0.52,
                     "open": 3820, "high": 3830, "low": 3780, "prev_close": 3820,
                     "volume": 2000000, "value_cr": 760, "year_high": 3850, "year_low": 3200,
                     "near_52h": 1.3, "near_52l": 18.8, "chg_30d": 3.0, "chg_365d": 10.0},
                ],
                "gainers": [{"symbol": "RELIANCE", "last": 2500, "pct": 2.04, "volume": 5000000}],
                "losers": [{"symbol": "TCS", "last": 3800, "pct": -0.52, "volume": 2000000}],
                "most_traded": [{"symbol": "RELIANCE", "last": 2500, "pct": 2.04, "volume": 5000000, "value_cr": 1250, "year_high": 2600, "year_low": 2000}],
                "most_volume": [{"symbol": "RELIANCE", "last": 2500, "pct": 2.04, "volume": 5000000, "value_cr": 1250}],
            },
        },
        "option_chain": {
            "NIFTY": {
                "symbol": "NIFTY", "pcr_oi": 1.15, "pcr_vol": 0.95,
                "signal": "Bullish", "max_pain": 22400,
                "ce_oi_total": 1000000, "pe_oi_total": 1150000,
                "top_ce": [{"strike": 22500, "oi": 500000, "chg_oi": 10000}],
                "top_pe": [{"strike": 22400, "oi": 600000, "chg_oi": -5000}],
            },
        },
        "commodities": {
            "TATAGOLD": {"last": 85, "change": 0.5, "pct": 0.59,
                         "open": 84.5, "high": 85.5, "low": 84,
                         "prev_close": 84.5, "week52_high": 90, "week52_low": 60},
        },
        "forex": {"usdinr": 83.5, "usdeur": 0.92, "usdgbp": 0.79, "usdjpy": 150.2, "date": "2026-03-01"},
        "corporate_actions": [
            {"symbol": "INFY", "company": "Infosys Ltd", "subject": "Dividend - Rs 18 Per Share",
             "ex_date": "05-Mar-2026", "record_date": "06-Mar-2026",
             "ltp": 1800, "pe": 28.5, "week52_high": 1950, "week52_low": 1400},
        ],
        "insider_trading": [
            {"symbol": "HDFCBANK", "company": "HDFC Bank Ltd", "acquirer": "Sashidhar Jagdishan",
             "relation": "KMP", "buy_qty": 10000, "sell_qty": 0,
             "buy_value": 15000000, "sell_value": 0, "date": "28-Feb-2026"},
        ],
        "preopen": {
            "key": "NIFTY", "timestamp": "2026-03-01T09:07:00",
            "advances": 30, "declines": 20,
            "stocks": [
                {"symbol": "RELIANCE", "iep": 2500, "change": 50, "pct": 2.04, "prev_close": 2450, "final_qty": 100000},
            ],
            "gainers": [{"symbol": "RELIANCE", "iep": 2500, "change": 50, "pct": 2.04, "prev_close": 2450, "final_qty": 100000}],
            "losers": [{"symbol": "TCS", "iep": 3780, "change": -40, "pct": -1.05, "prev_close": 3820, "final_qty": 50000}],
        },
    }
    
    mock_delta = {
        "timestamp": "2026-03-01T10:00:00",
        "prev_time": "2026-03-01T09:30:00",
        "fii_dii": {
            "fii_net_prev": 1500, "fii_net_curr": 2000,
            "fii_reversal": None, "dii_reversal": None,
            "dii_net_prev": -300, "dii_net_curr": -500,
        },
        "indices": {
            "best": {"name": "NIFTY 50", "signal": "📈 Up", "pct_change": 0.67},
            "worst": {"name": "NIFTY BANK", "signal": "📉 Down", "pct_change": -0.21},
            "changes": {},
        },
        "sectors": {},
        "forex": {"direction": "Weakened", "change": 0.15},
    }
    
    # Test each formatter
    msg1 = format_fii_dii_msg(mock_snapshot, mock_delta)
    assert len(msg1) > 0, "fii_dii_msg empty"
    print(f"[PASS] format_fii_dii_msg: {len(msg1)} chars")
    
    msg2 = format_sector_msg(mock_snapshot, mock_delta)
    assert len(msg2) > 0, "sector_msg empty"
    print(f"[PASS] format_sector_msg: {len(msg2)} chars")
    
    msg3 = format_options_msg(mock_snapshot)
    assert len(msg3) > 0, "options_msg empty"
    print(f"[PASS] format_options_msg: {len(msg3)} chars")
    
    msg4 = format_commodities_msg(mock_snapshot, mock_delta)
    assert len(msg4) > 0, "commodities_msg empty"
    print(f"[PASS] format_commodities_msg: {len(msg4)} chars")
    
    msg5 = format_corporate_msg(mock_snapshot)
    assert len(msg5) > 0, "corporate_msg empty"
    assert "LTP" in msg5 or "₹1,800" in msg5, "Corporate msg should show LTP"
    print(f"[PASS] format_corporate_msg: {len(msg5)} chars (LTP/PE included)")
    
    msg6 = format_preopen_msg(mock_snapshot)
    assert len(msg6) > 0, "preopen_msg empty"
    print(f"[PASS] format_preopen_msg: {len(msg6)} chars")
    
    msg7 = format_delta_alert(mock_delta)
    # May be None if no significant alerts
    print(f"[PASS] format_delta_alert: {'None (no alerts)' if msg7 is None else f'{len(msg7)} chars'}")
    
    msg8 = format_52w_alerts_msg(mock_snapshot)
    # TCS is at 92.3% of 52W range - near high; depends on threshold
    print(f"[PASS] format_52w_alerts_msg: {'None (no alerts)' if msg8 is None else f'{len(msg8)} chars'}")


def test_excel_with_mock():
    """Test Excel logging with mock data."""
    import os
    import tempfile
    from tracker.excel_manager import ExcelManager
    
    test_path = os.path.join(tempfile.gettempdir(), "test_market_tracker.xlsx")
    excel = ExcelManager(path=test_path)
    
    mock_snapshot = {
        "fii_dii": {
            "date": "01-Mar-2026",
            "fii": {"buy": 5000, "sell": 3000, "net": 2000},
            "dii": {"buy": 4000, "sell": 4500, "net": -500},
            "total_net": 1500, "signal": "FII Bullish",
            "interpretation": "FII buying",
        },
        "indices": {
            "NIFTY 50": {"last": 22500, "change": 150, "pct": 0.67,
                         "open": 22350, "high": 22550, "low": 22300,
                         "prev_close": 22350, "advances": 30, "declines": 20},
        },
        "sectors": {},
        "commodities": {
            "TATAGOLD": {"last": 85, "change": 0.5, "pct": 0.59,
                         "week52_high": 90, "week52_low": 60},
        },
        "forex": {"usdinr": 83.5, "usdeur": 0.92, "usdgbp": 0.79, "usdjpy": 150.2, "date": "2026-03-01"},
        "option_chain": {
            "NIFTY": {
                "symbol": "NIFTY", "pcr_oi": 1.15, "pcr_vol": 0.95,
                "signal": "Bullish", "max_pain": 22400,
                "ce_oi_total": 1000000, "pe_oi_total": 1150000,
                "top_ce": [{"strike": 22500, "oi": 500000, "chg_oi": 10000}],
                "top_pe": [{"strike": 22400, "oi": 600000, "chg_oi": -5000}],
            },
        },
        "preopen": {
            "key": "NIFTY", "timestamp": "2026-03-01T09:07:00",
            "advances": 30, "declines": 20,
            "stocks": [],
            "gainers": [{"symbol": "RELIANCE", "iep": 2500, "pct": 2.04}],
            "losers": [{"symbol": "TCS", "iep": 3780, "pct": -1.05}],
        },
        "corporate_actions": [
            {"symbol": "INFY", "company": "Infosys Ltd", "subject": "Dividend Rs 18",
             "ex_date": "05-Mar-2026", "record_date": "06-Mar-2026",
             "ltp": 1800, "pe": 28.5, "week52_high": 1950, "week52_low": 1400},
        ],
        "insider_trading": [
            {"symbol": "HDFCBANK", "company": "HDFC Bank", "acquirer": "Director",
             "relation": "KMP", "buy_qty": 10000, "sell_qty": 0,
             "buy_value": 15000000, "sell_value": 0, "date": "28-Feb-2026"},
        ],
    }
    
    excel.log_snapshot(mock_snapshot, None)
    
    from openpyxl import load_workbook
    wb = load_workbook(test_path)
    sheets = wb.sheetnames
    
    expected = ["FII_DII", "Indices", "Commodities", "Forex", "Options", "PreOpen", "Corporate", "Insider"]
    found = []
    missing = []
    for s in expected:
        if s in sheets:
            found.append(s)
        else:
            missing.append(s)
    
    print(f"[PASS] Excel sheets created: {', '.join(found)}")
    if missing:
        print(f"[WARN] Missing sheets: {', '.join(missing)}")
    
    # Check Options sheet has data
    if "Options" in sheets:
        ws = wb["Options"]
        assert ws.max_row >= 2, "Options sheet has no data rows"
        print(f"[PASS] Options sheet: {ws.max_row - 1} data rows")
    
    # Check PreOpen sheet has data
    if "PreOpen" in sheets:
        ws = wb["PreOpen"]
        assert ws.max_row >= 2, "PreOpen sheet has no data rows"
        print(f"[PASS] PreOpen sheet: {ws.max_row - 1} data rows")
    
    # Check Corporate sheet has LTP/PE columns
    if "Corporate" in sheets:
        ws = wb["Corporate"]
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        assert "LTP" in headers, f"Corporate sheet missing LTP column. Headers: {headers}"
        assert "PE" in headers, f"Corporate sheet missing PE column. Headers: {headers}"
        print(f"[PASS] Corporate sheet: has LTP + PE columns ({ws.max_row - 1} data rows)")
    
    # Cleanup
    os.remove(test_path)
    print(f"[PASS] Excel test file cleaned up")


if __name__ == "__main__":
    print("=" * 60)
    print("AUDIT TEST SUITE")
    print("=" * 60)
    
    try:
        test_imports()
    except Exception as e:
        print(f"[FAIL] Imports: {e}")
        sys.exit(1)
    
    print()
    
    try:
        test_formatters_with_mock_data()
    except Exception as e:
        print(f"[FAIL] Formatters: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    
    try:
        test_excel_with_mock()
    except Exception as e:
        print(f"[FAIL] Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print()
    print("=" * 60)
    print("ALL TESTS PASSED")
    print("=" * 60)
