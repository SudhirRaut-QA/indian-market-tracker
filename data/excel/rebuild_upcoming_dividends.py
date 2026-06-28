import argparse
import csv
import difflib
import re
import shutil
import sys
import time
from datetime import datetime, date
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]

if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from tracker.nse_scraper import MarketScraper
    HAS_TRACKER = True
except Exception:
    HAS_TRACKER = False
    MarketScraper = None

CSV_HEADERS = [
    "Company",
    "Symbol",
    "Exchange",
    "Ex_Date",
    "Record_Date",
    "Dividend_Rs",
    "LTP_INR",
    "Dividend_Yield_Percent",
    "PE_Ratio",
    "Buy_NotBuy_Indicator",
    "Buy_Zone_INR",
    "Sell_Target_INR",
    "Stop_Loss_INR",
    "Notes",
]

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

EXCLUDED_SYMBOLS = {"ANZEN", "CUBEINVIT", "MAPLEINVIT"}
ALLOWED_NON_EQUITY_SYMBOLS = {"CAPINVIT"}


def is_excluded_row(company: str, symbol: str) -> bool:
    s = (symbol or "").strip().upper()
    c = (company or "").strip().lower()

    if s in ALLOWED_NON_EQUITY_SYMBOLS:
        return False

    if s in EXCLUDED_SYMBOLS:
        return True

    # Trust/InvIT entries often behave like distributions and can mismatch broker dividend views.
    if "trust" in c or s.endswith("INVIT"):
        return True

    return False


def get_manual_overrides(today: date, days_ahead: int) -> list[dict]:
    cap_ex = date(2026, 5, 27)
    horizon = today.fromordinal(today.toordinal() + days_ahead)

    if cap_ex < today or cap_ex > horizon:
        return []

    ltp = 72.24
    div_rs = 2.40
    pe = 4.35
    div_yield = round((div_rs / ltp) * 100.0, 2)
    signal, buy_zone, sell_target, stop_loss = classify_signal(ltp, pe)

    return [
        {
            "Company": "Capital Infra Trust",
            "Symbol": "CAPINVIT",
            "Exchange": "NSE/BSE",
            "Ex_Date": fmt_date(cap_ex),
            "Record_Date": fmt_date(cap_ex),
            "Dividend_Rs": fmt_num(div_rs, 2),
            "LTP_INR": fmt_num(ltp, 2),
            "Dividend_Yield_Percent": fmt_num(div_yield, 2),
            "PE_Ratio": fmt_num(pe, 2),
            "Buy_NotBuy_Indicator": signal,
            "Buy_Zone_INR": buy_zone,
            "Sell_Target_INR": sell_target,
            "Stop_Loss_INR": stop_loss,
            "Notes": "Manual verified include: user-requested CAPINVIT (Dividend Rs 2.40, LTP near 72.24)",
        }
    ]


def merge_overrides(rows: list[dict], overrides: list[dict]) -> list[dict]:
    if not overrides:
        return rows

    idx = {(r.get("Symbol", "").upper(), r.get("Ex_Date", "")): i for i, r in enumerate(rows)}
    for o in overrides:
        key = ((o.get("Symbol") or "").upper(), o.get("Ex_Date") or "")
        if key in idx:
            rows[idx[key]] = o
        else:
            rows.append(o)

    rows.sort(key=lambda r: (r.get("Ex_Date", ""), r.get("Symbol", "")))
    return rows


def extract_dividend_amount(subject: str) -> float:
    patterns = [
        r"Rs\.?\s*([\d.]+)\s*(?:/|-|per)\s*(?:share|shr)",
        r"Re\.?\s*([\d.]+)\s*(?:/|-|per)\s*(?:share|shr)",
        r"Rs\.?\s*([\d.]+)",
        r"Re\.?\s*([\d.]+)",
    ]
    for pat in patterns:
        match = re.search(pat, subject or "", re.IGNORECASE)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                pass
    return 0.0


def parse_nse_date(text: str) -> date | None:
    if not text:
        return None
    for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def fmt_date(d: date | None) -> str:
    return d.isoformat() if d else ""


def fmt_num(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"


def parse_ddmmyyyy(text: str) -> date | None:
    if not text:
        return None
    try:
        return datetime.strptime(text, "%d/%m/%Y").date()
    except ValueError:
        return None


def parse_iso_date(text: str) -> date | None:
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_numeric(text: str) -> float | None:
    if text is None:
        return None
    raw = str(text).strip()
    if not raw:
        return None
    if raw.upper() in {"NA", "UNAVAILABLE", "NOT_APPLICABLE"}:
        return None
    try:
        return float(raw.replace(",", ""))
    except ValueError:
        return None


def enforce_quality(rows: list[dict], today: date, days_ahead: int) -> list[dict]:
    filtered: list[dict] = []
    horizon = today.fromordinal(today.toordinal() + days_ahead)

    for row in rows:
        symbol = (row.get("Symbol") or "").strip().upper()
        company = (row.get("Company") or "").strip()

        if is_excluded_row(company, symbol):
            continue

        ex_date = parse_iso_date((row.get("Ex_Date") or "").strip())
        if ex_date is None or ex_date < today or ex_date > horizon:
            continue

        dividend = parse_numeric(row.get("Dividend_Rs", ""))
        if dividend is None or dividend <= 0:
            continue

        ltp = parse_numeric(row.get("LTP_INR", ""))
        if ltp is None or ltp <= 0:
            continue

        filtered.append(row)

    filtered.sort(key=lambda r: (r.get("Ex_Date", ""), r.get("Symbol", "")))
    return filtered


def load_company_symbol_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    candidates = sorted(SCRIPT_DIR.glob("upcoming_dividends_full_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    latest_csv = SCRIPT_DIR / "upcoming_dividends_latest.csv"
    if latest_csv.exists():
        candidates.insert(0, latest_csv)
    for csv_path in candidates[:3]:
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = (row.get("Company") or "").strip()
                    symbol = (row.get("Symbol") or "").strip().upper()
                    if name and symbol and name.lower() not in mapping:
                        mapping[name.lower()] = symbol
        except Exception:
            continue
    return mapping


def load_seed_rows() -> list[dict]:
    seed_path = SCRIPT_DIR / "upcoming_dividends_seed.csv"
    if seed_path.exists():
        try:
            with seed_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                rows = []
                for row in reader:
                    rows.append({h: (row.get(h, "") or "") for h in CSV_HEADERS})
                if rows:
                    return rows
        except Exception:
            pass

    candidates = sorted(SCRIPT_DIR.glob("upcoming_dividends_full_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    for csv_path in candidates:
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                rows = []
                for row in reader:
                    rows.append({h: (row.get(h, "") or "") for h in CSV_HEADERS})
                if rows:
                    return rows
        except Exception:
            continue
    return []


def resolve_symbol(company: str, company_symbol_map: dict[str, str]) -> str:
    key = company.strip().lower()
    if key in company_symbol_map:
        return company_symbol_map[key]

    if company_symbol_map:
        matches = difflib.get_close_matches(key, list(company_symbol_map.keys()), n=1, cutoff=0.74)
        if matches:
            return company_symbol_map[matches[0]]

    cleaned = re.sub(r"[^A-Za-z0-9]", "", company).upper()
    return cleaned[:12]


def fetch_moneycontrol_upcoming() -> list[dict]:
    url = "https://www.moneycontrol.com/stocks/marketinfo/dividends_declared/index.php"
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=20)
    if resp.status_code != 200:
        return []

    text = re.sub(r"\s+", " ", resp.text)
    pattern = re.compile(
        r"\[(?P<company>[^\]]+?)\]\(#\)\s*"
        r"LTP(?P<ltp>[0-9,]+(?:\.[0-9]+)?)\([^)]*\)\s*"
        r"Dividend \(Rs\.\)(?P<div>[0-9]+(?:\.[0-9]+)?)\s*"
        r"Announcement Date(?P<ann>\d{2}/\d{2}/\d{4})\s*"
        r"Ex-Date(?P<ex>\d{2}/\d{2}/\d{4})",
        re.IGNORECASE,
    )

    rows = []
    for m in pattern.finditer(text):
        try:
            company = m.group("company").strip()
            ltp = float(m.group("ltp").replace(",", ""))
            div = float(m.group("div"))
            ex_date = parse_ddmmyyyy(m.group("ex"))
            if not company or not ex_date:
                continue
            rows.append(
                {
                    "company": company,
                    "ltp": ltp,
                    "dividend": div,
                    "ex_date": ex_date,
                    "record_date": ex_date,
                    "subject": f"Dividend - Rs {div} per share",
                    "source": "Moneycontrol dividends",
                }
            )
        except Exception:
            continue
    return rows


def build_from_seed(throttle: float) -> list[dict]:
    seed_rows = load_seed_rows()
    refreshed = []
    for row in seed_rows:
        if is_excluded_row(row.get("Company", ""), row.get("Symbol", "")):
            continue

        symbol = (row.get("Symbol") or "").strip().upper()
        if not symbol:
            continue

        ltp = None
        pe = None
        divy = None

        fallback_quote = parse_screener(symbol)
        if fallback_quote:
            ltp = fallback_quote.get("ltp")
            pe = fallback_quote.get("pe")
            divy = fallback_quote.get("divy")

        try:
            div_rs = float(row.get("Dividend_Rs", "0")) if row.get("Dividend_Rs", "").upper() != "NA" else 0.0
        except ValueError:
            div_rs = 0.0

        if divy is None and ltp and ltp > 0 and div_rs > 0:
            divy = round((div_rs / ltp) * 100.0, 2)

        signal, buy_zone, sell_target, stop_loss = classify_signal(ltp, pe)

        if ltp is not None:
            row["LTP_INR"] = fmt_num(ltp, 2)
            row["Buy_NotBuy_Indicator"] = signal
            row["Buy_Zone_INR"] = buy_zone
            row["Sell_Target_INR"] = sell_target
            row["Stop_Loss_INR"] = stop_loss
        if pe is not None and pe > 0:
            row["PE_Ratio"] = fmt_num(pe, 2)
        if divy is not None:
            row["Dividend_Yield_Percent"] = fmt_num(divy, 2)

        note = (row.get("Notes") or "").strip()
        source_note = "Universe fallback: previous file + live Screener quote refresh"
        row["Notes"] = f"{note}; {source_note}" if note else source_note

        refreshed.append(row)
        time.sleep(max(0.0, throttle))

    refreshed.sort(key=lambda r: (r.get("Ex_Date", ""), r.get("Symbol", "")))
    return refreshed


def parse_screener(symbol: str) -> dict | None:
    urls = (
        f"https://www.screener.in/company/{symbol}/",
        f"https://www.screener.in/company/{symbol}/consolidated/",
    )

    for url in urls:
        try:
            resp = requests.get(url, headers=HTTP_HEADERS, timeout=15)
            if resp.status_code != 200:
                continue
            text = resp.text

            ltp = None
            pe = None
            divy = None

            m = re.search(r"Current Price\s*₹\s*([0-9,]+(?:\.[0-9]+)?)", text)
            if m:
                ltp = float(m.group(1).replace(",", ""))

            m = re.search(r"Stock P/E\s*([0-9]+(?:\.[0-9]+)?)", text)
            if m:
                pe = float(m.group(1))

            m = re.search(r"Dividend Yield\s*([0-9]+(?:\.[0-9]+)?)\s*%", text)
            if m:
                divy = float(m.group(1))

            if ltp is None and pe is None and divy is None:
                continue

            return {
                "ltp": ltp,
                "pe": pe,
                "divy": divy,
                "source": "Screener (latest close)",
            }
        except Exception:
            continue

    return None


def classify_signal(ltp: float | None, pe: float | None) -> tuple[str, str, str, str]:
    if ltp is None or ltp <= 0:
        return "WATCH", "NA", "NA", "NA"

    if pe is None or pe <= 0:
        signal = "WATCH"
    elif pe < 15:
        signal = "ACCUMULATE"
    elif pe < 30:
        signal = "WATCH/ACCUMULATE"
    elif pe < 55:
        signal = "WATCH"
    else:
        signal = "NOT BUY"

    buy_lo = ltp * 0.98
    buy_hi = ltp
    sell_lo = ltp * 1.06
    sell_hi = ltp * 1.12
    stop = ltp * 0.96

    buy_zone = f"{buy_lo:.0f}-{buy_hi:.0f}"
    sell_target = f"{sell_lo:.0f}-{sell_hi:.0f}"
    stop_loss = f"{stop:.0f}"

    return signal, buy_zone, sell_target, stop_loss


def build_rows(days_ahead: int, throttle: float) -> list[dict]:
    scraper = MarketScraper() if HAS_TRACKER else None
    actions = []
    if scraper is not None:
        try:
            actions = scraper.get_corporate_actions(days_range=days_ahead) or []
        except Exception:
            actions = []
    company_symbol_map = load_company_symbol_map()

    today = date.today()
    rows: list[dict] = []
    seen = set()

    if not actions:
        fallback_rows = fetch_moneycontrol_upcoming()
        for item in fallback_rows:
            if item["ex_date"] < today:
                continue
            if item["ex_date"] > today.fromordinal(today.toordinal() + days_ahead):
                continue

            company = item["company"]
            symbol = resolve_symbol(company, company_symbol_map)
            if is_excluded_row(company, symbol):
                continue
            ltp = item["ltp"]
            div_rs = item["dividend"]
            ex_date = item["ex_date"]
            record_date = item["record_date"]

            pe = None
            div_yield = round((div_rs / ltp) * 100.0, 2) if ltp > 0 else None
            note_parts = ["Source: Moneycontrol dividends"]

            quote = None
            if scraper is not None:
                try:
                    quote = scraper.get_stock_quote(symbol)
                except Exception:
                    quote = None
            if quote:
                try:
                    q_ltp = quote.get("last")
                    q_pe = quote.get("pe")
                    if q_ltp not in (None, "", 0):
                        ltp = float(q_ltp)
                    if q_pe not in (None, ""):
                        pe = float(q_pe)
                    if div_rs > 0 and ltp > 0:
                        div_yield = round((div_rs / ltp) * 100.0, 2)
                    note_parts.append("NSE quote-equity")
                except (ValueError, TypeError):
                    pass

            if pe in (None, 0):
                fallback_quote = parse_screener(symbol)
                if fallback_quote:
                    if fallback_quote.get("ltp"):
                        ltp = fallback_quote["ltp"]
                    if fallback_quote.get("pe"):
                        pe = fallback_quote["pe"]
                    if fallback_quote.get("divy"):
                        div_yield = fallback_quote["divy"]
                    note_parts.append("Screener fallback")

            signal, buy_zone, sell_target, stop_loss = classify_signal(ltp, pe)

            rows.append(
                {
                    "Company": company,
                    "Symbol": symbol,
                    "Exchange": "NSE/BSE",
                    "Ex_Date": fmt_date(ex_date),
                    "Record_Date": fmt_date(record_date),
                    "Dividend_Rs": fmt_num(div_rs, 2) if div_rs > 0 else "NA",
                    "LTP_INR": fmt_num(ltp, 2) if ltp is not None else "UNAVAILABLE",
                    "Dividend_Yield_Percent": fmt_num(div_yield, 2) if div_yield is not None else "UNAVAILABLE",
                    "PE_Ratio": fmt_num(pe, 2) if pe is not None and pe > 0 else "UNAVAILABLE",
                    "Buy_NotBuy_Indicator": signal,
                    "Buy_Zone_INR": buy_zone,
                    "Sell_Target_INR": sell_target,
                    "Stop_Loss_INR": stop_loss,
                    "Notes": f"{item['subject']}; {'; '.join(note_parts)}",
                }
            )
            time.sleep(max(0.0, throttle))

        if rows:
            rows.sort(key=lambda r: (r["Ex_Date"], r["Symbol"]))
            merged = merge_overrides(rows, get_manual_overrides(today, days_ahead))
            return enforce_quality(merged, today, days_ahead)

        # Final fallback: use previously generated universe and refresh live quote fields.
        merged = merge_overrides(build_from_seed(throttle), get_manual_overrides(today, days_ahead))
        return enforce_quality(merged, today, days_ahead)

    for a in actions:
        subject = (a.get("subject") or "").strip()
        if "dividend" not in subject.lower():
            continue

        symbol = (a.get("symbol") or "").strip().upper()
        company = (a.get("company") or symbol).strip()
        if not symbol:
            symbol = resolve_symbol(company, company_symbol_map)
        if is_excluded_row(company, symbol):
            continue
        ex_date = parse_nse_date((a.get("ex_date") or "").strip())
        record_date = parse_nse_date((a.get("record_date") or "").strip())

        if ex_date and ex_date < today:
            continue

        key = (symbol, fmt_date(ex_date), subject)
        if key in seen:
            continue
        seen.add(key)

        div_rs = extract_dividend_amount(subject)

        ltp = None
        pe = None
        div_yield = None
        note_parts = ["Source: NSE corporate actions"]

        quote = None
        if scraper is not None:
            try:
                quote = scraper.get_stock_quote(symbol)
            except Exception:
                quote = None
        if quote:
            try:
                ltp_raw = quote.get("last")
                pe_raw = quote.get("pe")
                ltp = float(ltp_raw) if ltp_raw not in (None, "") else None
                pe = float(pe_raw) if pe_raw not in (None, "") else None
                note_parts.append("NSE quote-equity")
            except (ValueError, TypeError):
                pass

        if ltp is None:
            fallback = parse_screener(symbol)
            if fallback:
                ltp = fallback.get("ltp")
                pe = pe if pe not in (None, 0) else fallback.get("pe")
                div_yield = fallback.get("divy")
                note_parts.append("Screener fallback")

        if div_yield is None and div_rs > 0 and ltp and ltp > 0:
            div_yield = round((div_rs / ltp) * 100.0, 2)

        signal, buy_zone, sell_target, stop_loss = classify_signal(ltp, pe)

        row = {
            "Company": company,
            "Symbol": symbol,
            "Exchange": "NSE/BSE",
            "Ex_Date": fmt_date(ex_date),
            "Record_Date": fmt_date(record_date),
            "Dividend_Rs": fmt_num(div_rs, 2) if div_rs > 0 else "NA",
            "LTP_INR": fmt_num(ltp, 2) if ltp is not None else "UNAVAILABLE",
            "Dividend_Yield_Percent": fmt_num(div_yield, 2) if div_yield is not None else "UNAVAILABLE",
            "PE_Ratio": fmt_num(pe, 2) if pe is not None and pe > 0 else "UNAVAILABLE",
            "Buy_NotBuy_Indicator": signal,
            "Buy_Zone_INR": buy_zone,
            "Sell_Target_INR": sell_target,
            "Stop_Loss_INR": stop_loss,
            "Notes": f"{subject}; {'; '.join(note_parts)}",
        }

        rows.append(row)
        time.sleep(max(0.0, throttle))

    if not rows:
        merged = merge_overrides(build_from_seed(throttle), get_manual_overrides(today, days_ahead))
        return enforce_quality(merged, today, days_ahead)

    rows.sort(key=lambda r: (r["Ex_Date"], r["Symbol"]))
    merged = merge_overrides(rows, get_manual_overrides(today, days_ahead))
    return enforce_quality(merged, today, days_ahead)


def write_csv(csv_path: Path, rows: list[dict]) -> None:
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(xlsx_path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Upcoming_Dividends"

    ws.append(CSV_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in CSV_HEADERS])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(CSV_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col_name in enumerate(CSV_HEADERS, start=1):
        max_len = len(col_name)
        for r in range(2, ws.max_row + 1):
            value = ws.cell(row=r, column=idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild upcoming dividend list from scratch")
    parser.add_argument("--days-ahead", type=int, default=45, help="Days ahead to include ex-dates")
    parser.add_argument("--throttle", type=float, default=0.2, help="Delay between quote requests")
    parser.add_argument(
        "--keep-dated-snapshot",
        action="store_true",
        help="Also write dated CSV/XLSX snapshots in addition to latest files",
    )
    args = parser.parse_args()

    today_text = date.today().isoformat()
    dated_csv = SCRIPT_DIR / f"upcoming_dividends_full_{today_text}.csv"
    dated_xlsx = SCRIPT_DIR / f"upcoming_dividends_full_{today_text}.xlsx"
    latest_csv = SCRIPT_DIR / "upcoming_dividends_latest.csv"
    latest_xlsx = SCRIPT_DIR / "upcoming_dividends_latest.xlsx"

    rows = build_rows(days_ahead=args.days_ahead, throttle=args.throttle)
    write_csv(latest_csv, rows)
    write_xlsx(latest_xlsx, rows)

    if args.keep_dated_snapshot:
        shutil.copyfile(latest_csv, dated_csv)
        shutil.copyfile(latest_xlsx, dated_xlsx)

    print(f"REBUILT_ROWS:{len(rows)}")
    print(f"LATEST_CSV:{latest_csv}")
    print(f"LATEST_XLSX:{latest_xlsx}")
    if args.keep_dated_snapshot:
        print(f"DATED_CSV:{dated_csv}")
        print(f"DATED_XLSX:{dated_xlsx}")


if __name__ == "__main__":
    main()
